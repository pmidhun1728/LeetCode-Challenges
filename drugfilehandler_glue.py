import sys
import datetime
from datetime import datetime
from awsglue.utils import getResolvedOptions
import boto3
import uuid
import psycopg2
import base64
from email.mime.application import MIMEApplication
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

args = getResolvedOptions(sys.argv, ["file_name","region","bucket_name","receiverEmails","senderEmail","returnPathArn","sourceArn","dbhost","dbname","username","password","deltaReportMaxLimitMB"])


bucket_name = args['bucket_name']
region = args['region']
receiverEmails = args['receiverEmails']
senderEmail = args['senderEmail']
returnPathArn = args['returnPathArn']
sourceArn = args['sourceArn']
delta_report_max_limit_mb = int(args['deltaReportMaxLimitMB'])

# PostgreSQL connection details
db_host =  args['dbhost']
db_name = args['dbname']
db_user = args['username']
db_password = args['password']

# S3 client - global
s3client = None

CVS_S3_PREFIX = "drugfile/CVS/"
DELTA_CVS_VS_CVS_S3_PREFIX = "delta/CVS/CVS_vs_CVS/"

def main():
    print("[MAIN] Glue job started at:", datetime.now())
    global status, s3client

    print(f"[MAIN] - file_name: {args['file_name']}")
    print(f"[MAIN] - bucket_name: {args['bucket_name']}")
    print(f"[MAIN] - receiverEmails: {args['receiverEmails']}")
    print(f"[MAIN] - dbhost: {args['dbhost']}")

try:
    file_path = args['file_name']
    

    s3client = boto3.client("s3")
    s3Object = s3client.get_object(Bucket=bucket_name, Key=file_path)
    file_content = s3Object['Body'].read().decode('utf-8')

    file_content_for_archive = file_content

    if(file_content != None):
        status = "processed"
    

    lines = file_content.split('\n')
   
    file_name = file_path[-36:-4]
    print(file_name)

    timestampStr = file_path[-19:-4].replace("_","")

    month = int(timestampStr[0:2])
    day = int(timestampStr[2:4])
    year = int(timestampStr[4:8])
    hour = int(timestampStr[8:10])
    minute = int(timestampStr[10:12])
    second = int(timestampStr[12:14])

    print("it is coming after file")

    timestampstring = str(year)+str("-")+str(month)+str("-")+str(day)+str(" ")+str(hour)+str(":")+str(minute)+str(":")+str(second)

    timestamp_format = "%Y-%m-%d %H:%M:%S"
    timestamp_datetime = datetime.strptime(timestampstring, timestamp_format)

    lines = lines[1:-1]
    data_to_insert = []

    dataValidation(lines, file_name, timestamp_datetime)

    #Default  Value for create name and update name
    default_value_for_Column_one = "SYSTEM"

    print("looping the data")

    for line in lines:
        values = line.strip().split('|')
        if len(values) >= 8:
            data_row = (str(uuid.uuid4()),
                        file_name, values[0],
                        None if values[1] == '' else values[1],
                        None if values[2] == '' else values[2],
                        None if values[3] == '' else values[3],
                        None if values[4] == '' else values[4],
                        None if values[5] == '' else values[5],
                        None if values[6] == '' else values[6],
                        values[7],
                        datetime.now(),
                        default_value_for_Column_one)

            data_to_insert.append(data_row)

    print("calling persitData function to persist data in database")

    persistData(data_to_insert, file_content_for_archive, file_name)

    cvs_vs_cvs_delta = generate_cvs_to_cvs_delta_report(s3client)
    print("CVS vs CVS Delta report generated")

    cvs_vs_feops_delta = generate_cvs_to_feops_delta_report(s3client)
    print("CVS vs FeOps Delta report generated")

    # Get baseline_exists from cvs_vs_cvs_delta (or default to True if not available)
    baseline_exists = cvs_vs_cvs_delta.get('baseline_exists', True) if cvs_vs_cvs_delta else True

    print("data persisted now invoking procedure")

    print("End OF execution")

except Exception as e:
            error_message = str(e)
            print(f"Glue job failed with error: {error_message}")
            import traceback
            traceback_str = traceback.format_exc()
            print(f"Traceback: {traceback_str}")

            send_glue_job_failure_email(file_name, error_message)
            raise

def dataValidation(datas, file_name, timestamp_datetime):

    for data in datas:
        values = data.strip().split('|')
        if len(values) >= 8:
            if(len(values[0])<11 or len(values[0])>15 or values[0]==""):
                sendEmail(file_name, "NDC is Required Field and it's length Shlouldnot be less than 11 and shouldnot be more than 15", "NDC", timestamp_datetime, values[0])
            if(values[1]==""):
                sendEmail(file_name, "Description  is Required Field", "Description", timestamp_datetime, values[1])
            if(len(values[4])<14 or len(values[4])>14 or values[4]==""):
                sendEmail(file_name, "GPI is Required Field and it should be of 14 length", "GPI", timestamp_datetime, values[4])
            if(len(values[7])==11 or values[7]==""):
                sendEmail(file_name, "Preference Rank is Required Field", "Preference Rank", timestamp_datetime, values[7])

    persistDataInLogTable(file_name, 'Passed')

def sendEmail(file_name, exception, column_type, timestamp_datetime, values):

    print("Coming to send mail")
    print(file_name)
    print(exception)
    print(column_type)
    print(values)

    body = ""
    body += "Failed to Validate DrugFile"+"<br><br>"
    body += "<b>Filename: </b>" + file_name+"<br>"
    body += "<b>Exception: </b>" + exception+"<br>"
    body += "<b>Description: </b>" + column_type+"<br>"
    body += "<b>Received Date: </b>" + str(timestamp_datetime)

    print(body)
    email_list = receiverEmails.split(',')

    ses = boto3.client('ses', region_name=region)

    response = ses.send_email(
            Destination={
                'ToAddresses': email_list,
            },
            Message={
                'Body': {
                    'Html': {
                        'Data': body,
                    },
                },
                'Subject': {
                    'Data': 'Drug File - Has Data Error',
                },
            },
            ReturnPathArn=returnPathArn,
            SourceArn=sourceArn,
            Source=senderEmail
        )
    print("Email sent")

    status = "failed"
    persistDataInLogTable(file_name, status)
    if(status == 'failed'):
        exit()



def persistDataInLogTable(file_name, status):
    try:
        conn = psycopg2.connect(
            host=db_host,
            port=5432,
            database=db_name,
            user=db_user,
            password=db_password,
            sslmode='require'
        )
    
        print("On Validation Failure Pushing data to Log Table")
        print(status)
        cursor = conn.cursor()

        data_row = (str(uuid.uuid4()),
                    file_name,
                    status,
                    datetime.now(),
                    0,0,0,0,0,0,0,0,0,
                    "N/A",
                    datetime.now(),
                    "SYSTEM",
                    datetime.now(),
                    "SYSTEM")

        insert_query = "INSERT INTO intelengine.ie_inboud_file_prcs_log(ie_inbnd_file_prcs_log_id, file_nm, prcs_stts_cd, prcs_dtm, totl_rcrd_cnt, hdr_rcrd_cnt, ftr_rcrd_cnt, dtl_rcrd_cnt, err_rcrd_cnt, add_rcrd_cnt, no_chng_rcrd_cnt, chng_rcrd_cnt, del_rcrd_cnt, err_desc, creat_dtm, creat_nm, updt_dtm, updt_nm) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        cursor.execute(insert_query, data_row)
        conn.commit()

    except Exception as e:
        print("Error:", e)
        conn.rollback()
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()




def persistData(data_to_insert, file_content_for_archive, file_name):

    try:
        conn = psycopg2.connect(
            host=db_host,
            port=5432,
            database=db_name,
            user=db_user,
            password=db_password,
            sslmode='require'
        )

        if conn is not None and not conn.closed:
            print("Connection has established with databse")
        else:
            print("Connection is closed")

        cursor = conn.cursor()

        delete_query = f"DELETE From intelengine.ie_inbnd_drug_file_stg"

        cursor.execute(delete_query)

        print("executing before insert query")

        insert_query = "INSERT INTO intelengine.ie_inbnd_drug_file_stg(ie_inbnd_drug_file_stg_id,file_nm, ndc, description, mdspn_packsize, uom, gpi, packsize, mail_dspnsg_unit, prfrnc_rank, creat_dtm, creat_nm) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

        for data_row in data_to_insert:
            cursor.execute(insert_query, data_row)
        conn.commit()

        invokeProcedure(file_content_for_archive, file_name)

    except Exception as e:
        print("Error:", e)
        persistDataInLogTable(file_name, 'failed')
        conn.rollback()
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

    print("successfully deleted the data and saved the file data to database")


def invokeProcedure(file_content_for_archive, file_name):
    try:
        conn = psycopg2.connect(
            host=db_host,
            port=5432,
            database=db_name,
            user=db_user,
            password=db_password,
            sslmode='require'
        )

        print("coming to stored procedure")

        print(file_name)
        cursor = conn.cursor()
        outputparam = ''

        cursor.execute('CALL intelengine.ie_drug_file_load(%s, %s)', (file_name, outputparam))
        outputparam =cursor.fetchone()[0]
        conn.commit()

        print("data of output param type:", outputparam)

        if(outputparam == 'Succeeded'):
            s3clienttodest = boto3.client('s3')
            try:
                response=s3clienttodest.put_object(Body=file_content_for_archive, Bucket=bucket_name, Key="archive/"+file_name)
                print("File uploded to archive and now removing from Drugfile")
                if response['ResponseMetadata']['HTTPStatusCode'] == 200:
                    s3clienttodest.delete_object(Bucket=bucket_name, Key="drugfile/"+file_name+".txt")
                    print("File Removed From Drug Folder")
                    sendEmail_with_attachment(file_content_for_archive,file_name)
                print("File uploaded Successfully in Archive Folder.")                             
            except Exception as e:
                print(e.__cause__)
                print("Error in uploading the file")
       
    except Exception as e:
        print("Error:", e)
        persistDataInLogTable(file_name, 'failed')
        conn.rollback()
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()

def sendEmail_with_attachment(file_content_for_archive,file_name):
    print("Sending drug file Mail")
    current_date_time = datetime.now()
    print("curent date is: ",current_date_time)
    current_date_formatted = current_date_time.strftime("%d-%m-%Y")
    print("formated date is: ",current_date_formatted)
    ses = boto3.client('ses', region_name=region)

    body = ""
    
    body += "<b>Please find the attached DP CVS Drug File</b>"+"<br><br>"
    body += "<b>Filename: </b>" + file_name+"<br>"
    body += "<b>Received Date: </b>" + str(current_date_formatted)+"<br><br>"
    body += "Thanks & Regards,"+"<br>"
    body += "CarelonRx"
    print("printing after body creation")
    
    msg = MIMEMultipart()
    print("printing after multipart")
    msg['Subject'] = 'CVS Drug File'
    # msg['From'] = senderEmail
    # msg['To'] = 'praveen.kumar5@carelon.com'
    print("printing after mail msg declared succesfully")

    msg.attach(MIMEText(body, 'html'))

    attachment = MIMEApplication(file_content_for_archive)
    attachment.add_header('Content-Disposition', 'attachment', filename=file_name)
    msg.attach(attachment)

    print("printing after mail msg declared succesfully")

    # file_content_for_archive['Content-Disposition'] = f'attachment; filename="yuorfile.txt"'
    # msg.attach(file_content_for_archive)
    print("printing after msg attach succesfully")
    raw_message = msg.as_string()
    
    print("now sending response mail to participate")
    email_list = receiverEmails.split(',')
    # formatted_emails = ','.join([f"'{email}'" for email in email_list])
    # formatted_email_list = formatted_emails.split(',')
    print(email_list)
    response = ses.send_raw_email(
          Source = senderEmail,
          Destinations=email_list,
          RawMessage={'Data':raw_message}
    )
    print("Drug File Mail Has Been Sent")
    print(response)

def generate_cvs_to_feops_delta_report(s3client):
    """Generate delta report: NDCs in CVS but missing from FeOps."""
    file_path = args['file_name']
    
    # Extract filename and construct correct S3 path using CVS prefix
    if "/" in file_path:
        filename_only = file_path.split("/")[-1]
        correct_file_path = CVS_S3_PREFIX + filename_only
    else:
        correct_file_path = CVS_S3_PREFIX + file_path
    
    print(f"Reading CVS file from: {correct_file_path}")
    s3Object = s3client.get_object(Bucket=bucket_name, Key=correct_file_path)
    file_content = s3Object['Body'].read().decode('utf-8')
    cvs_df = pd.read_csv(BytesIO(file_content.encode('utf-8')), sep='|', dtype=str).fillna("")
    print(f"CVS file loaded with {len(cvs_df)} rows")

    # Filter out EOF rows
    key_col = "NDC"
    cvs_df = cvs_df[cvs_df[key_col] != 'EOF']

    # Read the FeOps file (Excel) from feops/ prefix
    feops_prefix = "feops/"
    feops_df = pd.DataFrame()

    try:
        response = s3client.list_objects_v2(Bucket=bucket_name, Prefix=feops_prefix)

        if 'Contents' in response:
            excel_files = [obj for obj in response['Contents']
                           if (obj['Key'].endswith('.xlsx') or obj['Key'].endswith('.xls'))
                           and not obj['Key'].endswith("/")
                           and obj['Key'].count('/') == 1  # Only files directly in feops/, not subfolders
                           and not obj['Key'].split('/')[-1].startswith('~$')]  # Exclude Excel temp files

            if excel_files:
                # Get the most recent FeOps file
                excel_files.sort(key=lambda x: x['LastModified'], reverse=True)
                feops_file_key = excel_files[0]['Key']
                print(f"Reading FeOps file: {feops_file_key}")

                feops_s3Object = s3client.get_object(Bucket=bucket_name, Key=feops_file_key)
                feops_df = pd.read_excel(BytesIO(feops_s3Object['Body'].read()), dtype=str).fillna("")
                print(f"FeOps file loaded with {len(feops_df)} rows")
            else:
                print("No FeOps Excel file found in the bucket")
                return
    except Exception as e:
        print(f"Error reading FeOps file: {e}")
        return

    # Ensure the key column exists in both DataFrames
    if key_col not in cvs_df.columns or key_col not in feops_df.columns:
        print(f"Column '{key_col}' not found in both files")
        return

    # Find NDCs present in CVS but missing from FeOps
    # Normalize NDCs to 11 digits (CVS has 15 digits, FeOps has 11 digits)
    cvs_ndcs = set(cvs_df[key_col].astype(str).str[-11:])
    feops_ndcs = set(feops_df[key_col].astype(str).str.zfill(11))
    missing_ndcs = cvs_ndcs - feops_ndcs

    print(f"Missing NDC count (CVS vs FeOps): {len(missing_ndcs)}")

    # Create delta DataFrame
    delta_df = pd.DataFrame({"NDC": list(missing_ndcs)})

    # Output file
    timestamp = datetime.now().strftime("%m%d%Y_%H%M%S")
    file_name = f"Delta_CVS_vs_SubList_{timestamp}.txt"
    output_key = f"delta/CVS/CVS_vs_FeOps/{file_name}"

    file_content = None

    if not delta_df.empty:
        # Save as text file
        buffer = BytesIO()
        delta_df.to_csv(buffer, index=False)
        file_content = buffer.getvalue()
        s3client.put_object(Bucket=bucket_name, Key=output_key, Body=file_content)
        print(f"CVS vs FeOps delta report uploaded to s3://{bucket_name}/{output_key}. Missing NDCs: {len(missing_ndcs)}")
    else:
        print("No differences found between CVS and FeOps files.")

    return {'file_content': file_content, 'file_name': file_name}


def parse_filename_timestamp(filename):
    """
    Parse timestamp from filename format: DP_CVS_Drug_File_MMDDYYYY_HHMMSS.txt
    Returns datetime object or datetime.min if parsing fails
    """
    try:
        # Extract timestamp from filename: DP_CVS_Drug_File_01212026_030309.txt -> 01212026_030309
        timestamp_str = "_".join(filename.split("_")[-2:]).replace(".txt", "")
        
        # Parse the timestamp: MMDDYYYY_HHMMSS
        parsed_datetime = datetime.strptime(timestamp_str, "%m%d%Y_%H%M%S")
        return parsed_datetime
        
    except ValueError:
        # Try alternative date format (DDMMYYYY instead of MMDDYYYY)
        try:
            parts = timestamp_str.split('_')
            if len(parts) == 2 and len(parts[0]) == 8 and len(parts[1]) == 6:
                alt_timestamp = f"{parts[0][:2]}{parts[0][2:4]}{parts[0][4:8]}_{parts[1]}"
                return datetime.strptime(alt_timestamp, "%d%m%Y_%H%M%S")
        except:
            pass
            
    except Exception as e:
        print(f"[BASELINE] Error parsing timestamp from {filename}: {e}")
    
    return datetime.min


def get_baseline_file(s3client,file_path):
    """
    get second latest  file from S3 bucket based on timestamp in the file name.
    Args:
        prefix: S3 prefix/folder path
    Returns:
        second latest file name
    """
    print(f"[FILE_READ] Searching for baseline file in bucket: {bucket_name}, prefix: {CVS_S3_PREFIX}")
    response = s3client.list_objects_v2(Bucket=bucket_name, Prefix=CVS_S3_PREFIX)

    if 'Contents' not in response:
        print(f"No files found in : {bucket_name}/{CVS_S3_PREFIX}")
        return ""

    # Filter text files and exclude the current file (compare by filename only)
    current_filename = file_path.split('/')[-1] if '/' in file_path else file_path
    text_files = [obj for obj in response['Contents']
                  if obj['Key'].endswith('.txt') and obj['Key'].split('/')[-1] != current_filename]


    # Sort files by timestamp in filename (format: <filename>_mmddYYYY_HHMMSS.txt) in descending order
    def extract_timestamp(file_obj):
        filename = file_obj['Key'].split('/')[-1]  # Get filename from full S3 key
        return parse_filename_timestamp(filename)

    sorted_files = sorted(text_files, key=extract_timestamp, reverse=True)

    # Get the most recent file from the filtered list (current file already excluded)
    if len(sorted_files) >= 1:
        baseline_file = sorted_files[0]  # Most recent after excluding current file
        baseline_file_key = baseline_file['Key']
        return baseline_file_key
    else:
        return ""



def read_latest_and_baseline_files(file_path, s3client):
    """
    Read the latest file and find the second most recent file as baseline
    Returns: tuple (latest_file_df, baseline_file_df)
    """
    # Read the current/latest file
    print(f"[FILE_READ] Starting file comparison - Latest: {file_path}")
    
    # Extract just the filename if file_path contains directory structure
    if "/" in file_path:
        filename_only = file_path.split("/")[-1]
        # Construct the correct S3 path using CVS prefix
        correct_file_path = CVS_S3_PREFIX + filename_only
        print(f"[FILE_READ] Extracted filename: {filename_only}")
        print(f"[FILE_READ] Constructed correct path: {correct_file_path}")
    else:
        correct_file_path = CVS_S3_PREFIX + file_path
        print(f"[FILE_READ] Using CVS prefix with filename: {correct_file_path}")
    
    try:
        print(f"[FILE_READ] Reading latest file from: {correct_file_path}")
        s3Object = s3client.get_object(Bucket=bucket_name, Key=correct_file_path)
        file_content = s3Object['Body'].read().decode('utf-8')
        latest_file_df = pd.read_csv(BytesIO(file_content.encode('utf-8')), sep='|', dtype=str).fillna("")
        print(f"[FILE_READ] Latest file loaded with {len(latest_file_df)} rows")
    except Exception as e:
        print(f"[FILE_READ] ERROR: Failed to read latest file {file_path}: {str(e)}")
        raise

    # Find the baseline file (second most recent file excluding the current one)
    baseline_file_df = pd.DataFrame()
    try:
        # Get the most recent file (baseline) - pass the filename only
        if "/" in file_path:
            filename_only = file_path.split("/")[-1]
        else:
            filename_only = file_path
        baseline_file_key = get_baseline_file(s3client, filename_only)
        print(f"Reading baseline file: {baseline_file_key}")

        # Read the baseline file as pipe-delimited CSV
        baseline_s3Object = s3client.get_object(Bucket=bucket_name, Key=baseline_file_key)
        baseline_file_content = baseline_s3Object['Body'].read().decode('utf-8')
        baseline_file_df = pd.read_csv(BytesIO(baseline_file_content.encode('utf-8')), sep='|', dtype=str).fillna("")
        print(f"Baseline file loaded with {len(baseline_file_df)} rows")

    except Exception as e:
        print(f"Error reading baseline file: {e}")
        baseline_file_df = pd.DataFrame()
    
    return latest_file_df, baseline_file_df

def generate_cvs_to_cvs_delta_report(s3client):
    file_path = args['file_name']
    print(f"[DELTA_REPORT] Starting CVS vs CVS delta report for file: {file_path}")
    
    # Read latest and baseline files
    latest_file_df, baseline_file_df = read_latest_and_baseline_files(file_path, s3client)
    key_col = "NDC"

    # Ensure the compare column exists in latest file (baseline can be empty)
    if key_col not in latest_file_df.columns:
        print(f"[DELTA_REPORT] ERROR: Column '{key_col}' not found in latest file")
        print(f"[DELTA_REPORT] Available columns in latest file: {list(latest_file_df.columns)}")
        raise ValueError(f"Column '{key_col}' not found in latest file")
    
    # Check baseline file only if it's not empty
    if not baseline_file_df.empty and key_col not in baseline_file_df.columns:
        print(f"[DELTA_REPORT] ERROR: Column '{key_col}' not found in baseline file")
        print(f"[DELTA_REPORT] Available columns in baseline file: {list(baseline_file_df.columns)}")
        raise ValueError(f"Column '{key_col}' not found in baseline file")

    # Filter out rows where NDC is 'EOF'
    latest_file_df = latest_file_df[latest_file_df[key_col] != 'EOF']
    if not baseline_file_df.empty:
        baseline_file_df = baseline_file_df[baseline_file_df[key_col] != 'EOF']

    # Find added, removed rows
    latest_file_df_NDCs = set(latest_file_df[key_col])
    if not baseline_file_df.empty:
        baseline_file_df_NDCs = set(baseline_file_df[key_col])
    else:
        baseline_file_df_NDCs = set()
        print(f"[DELTA_REPORT] No baseline file found - all records will be marked as additions")

    # Get added and removed rows
    added_NDCs = latest_file_df_NDCs - baseline_file_df_NDCs
    removed_NDCs = baseline_file_df_NDCs - latest_file_df_NDCs

    # Create DataFrames for added and removed rows
    added_rows = latest_file_df[latest_file_df[key_col].isin(added_NDCs)]
    removed_rows = baseline_file_df[baseline_file_df[key_col].isin(removed_NDCs)]

    # Create delta DataFrame
    delta_data = []

    # Add added rows
    for _, row in added_rows.iterrows():
        delta_data.append({
            'NDC': row[key_col],
            'Changes Detected': 'Addition'
        })

    # Add removed rows
    for _, row in removed_rows.iterrows():
        delta_data.append({
            'NDC': row[key_col],
            'Changes Detected': 'Deletion'
        })
    # Create delta DataFrame
    delta_df = pd.DataFrame(delta_data)

    # Output file
    timestamp = datetime.now().strftime("%m%d%Y_%H%M%S")
    file_name = f"Delta_cvs_vs_cvs_{timestamp}.xlsx"
    output_key = f"{DELTA_CVS_VS_CVS_S3_PREFIX}/{file_name}"

    baseline_exists = not baseline_file_df.empty
    file_content = None

    if not delta_df.empty:
        file_content = save_formatted_excel_to_s3(delta_df, bucket_name, output_key)
        print(f"Changes detected: {len(delta_df)} (Added: {len(added_rows)}, Removed: {len(removed_rows)})")
    else:
        print("No differences found between the files.")

    return {'file_content': file_content, 'file_name': file_name, 'baseline_exists': baseline_exists}



def save_formatted_excel_to_s3(df, bucket_name, output_key):
    """
    Save DataFrame to S3 as formatted Excel with styling
    """
    wb = format_excel_workbook(df)
    return upload_workbook_to_s3(wb, bucket_name, output_key)


def format_excel_workbook(df, worksheet_title="Drug File"):
    """
    Create and format an Excel workbook from DataFrame:
    - Auto-sized columns based on content
    - Bold headers with blue background
    """
    # Create a new workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = worksheet_title

    # Add data to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F4C75", end_color="0F4C75", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Apply header formatting
    for cell in ws[1]:  # First row (headers)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Auto-size columns based on content
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # Set column width with some padding
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

    return wb



def upload_workbook_to_s3(workbook, bucket_name, output_key):
    """
    Upload an Excel workbook to S3
    """
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    file_bytes = excel_buffer.getvalue()
    s3client.put_object(Bucket=bucket_name, Key=output_key, Body=file_bytes)
    print(f"Formatted delta file uploaded to S3: s3://{bucket_name}/{output_key}")
    return file_bytes

# -----------------------------------
# Email Triggering logic added for the CVS Delta Reports and S3 Cleanup

CVS_S3_PREFIX = "drugfile/CVS/"
DELTA_CVS_VS_CVS_S3_PREFIX = "delta/CVS/CVS_vs_CVS"

def send_cvs_delta_report_email(run_date, run_timestamp, dp_file, baseline_file, feops_file,
                                delta_cvs_vs_cvs_file, add_count, del_count, total_count,
                                delta_cvs_vs_feops_file, missing_ndc_count, to_emails,
                                cvs_vs_cvs_delta, cvs_vs_feops_delta, baseline_exists):
    """
    Send CVS delta report email with optional attachments.
    
    Args:
        cvs_vs_cvs_delta: dict with 'file_content' (bytes) and 'file_name' for CVS vs CVS delta
        cvs_vs_feops_delta: dict with 'file_content' (bytes) and 'file_name' for CVS vs FeOps delta
        baseline_exists: True if baseline file exists, False for first run
    
    Attachment rules:
        - Attach only if baseline_exists is True
        - Attach only if file size > 0 MB and <= DELTA_REPORT_MAX_LIMIT_MB
    """
    subject = f"CVS Delta Reports - {run_date} - {run_timestamp}"
    body = f"""
    CVS Daily Delta Reports Generated
    Run Date         : {run_date}
    DP               : CVS
    Inbound Drug File: {dp_file}
    Baseline Drug File: {baseline_file}
    FE Ops Sub List  : {feops_file}

    Report 1: CVS vs CVS (Additions and Deletions)
    Delta File Name   : {delta_cvs_vs_cvs_file}
    Summary (CVS vs CVS)
    ADDITION          : {add_count}
    DELETION          : {del_count}
    TOTAL IN REPORT   : {total_count}

    Report 2: CVS vs FE Ops (Missing NDCs)
    Delta File Name   : {delta_cvs_vs_feops_file}
    MISSING NDC COUNT : {missing_ndc_count}

    Notes:
    - CVS vs CVS report contains only NDCs added or deleted compared to baseline.
    - CVS vs FE Ops report contains NDCs present in today CVS file but missing in FE Ops substitution list.

    Thanks,
    Intel Engine
    """
    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg['From'] = senderEmail
    msg['To'] = ', '.join(to_emails)
    msg.attach(MIMEText(body, 'plain'))
    
    # Helper to attach delta file if conditions are met
    def attach_delta_file(delta_dict, report_name):
        if not delta_dict or not delta_dict.get('file_content'):
            return
        if not baseline_exists:
            print(f"Skipped {report_name} attachment (baseline missing - first run)")
            return
        file_size_mb = len(delta_dict['file_content']) / (1024 * 1024)
        if file_size_mb == 0:
            print(f"Skipped {report_name} attachment (file is empty)")
        elif file_size_mb > delta_report_max_limit_mb:
            print(f"Skipped {report_name} attachment (size {file_size_mb:.2f} MB > {delta_report_max_limit_mb} MB limit)")
        else:
            attachment = MIMEApplication(delta_dict['file_content'])
            attachment.add_header('Content-Disposition', 'attachment', filename=delta_dict['file_name'])
            msg.attach(attachment)
            print(f"Attached {report_name} delta: {delta_dict['file_name']} ({file_size_mb:.2f} MB)")
    
    attach_delta_file(cvs_vs_cvs_delta, "CVS vs CVS")
    attach_delta_file(cvs_vs_feops_delta, "CVS vs FeOps")
    
    ses = boto3.client('ses', region_name=region)
    response = ses.send_raw_email(
        Source=senderEmail,
        Destinations=to_emails,
        RawMessage={'Data': msg.as_string()}
    )
    print("Delta report email sent.")
    print(response)

def send_glue_job_failure_email(inbound_file, error_message):
    """
    Send failure notification email when Glue job fails before delta report generation.
    """
    subject = f"CVS Delta Reports - FAILED - {run_date} - {run_timestamp}"

    body = (
        "<b>ORx Delta Report Generation Failed</b><br><br>"
        f"<b>Run Date          :</b> {run_date}<br>"
        "<b>DP                :</b> CVS <br>"
        f"<b>Inbound Drug File :</b> {inbound_file}<br><br>"
        "<b>Status:</b><br>"
        "- Glue job failed before delta report generation completed.<br><br>"
        "<b>Error Summary:</b><br>"
        f"{error_message}<br><br>"
        "<b>Next Steps:</b><br>"
        "- Review Glue logs in CloudWatch.<br>"
        "- Re-run job after resolving issue.<br>"
        "- Inbound file remains in drugfile/ for retry (no cleanup performed).<br><br>"
        "<b>Thanks,</b><br>"
        "Intel Engine<br>"
    )

    send_email(
        subject=subject,
        body=body,
        receiver_emails=receiverEmails,
        region=region,
        return_path_arn=returnPathArn,
        source_arn=sourceArn,
        sender_email=senderEmail
    )
    print(f"Failure notification email sent for file: {inbound_file}")


if __name__ == "__main__":
    main()
